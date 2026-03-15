from io import BytesIO
from numbers import Number

import pandas as pd
from matplotlib import pyplot as plt
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet

from skylines.common.state import get_state


class Excel:

    def __init__(self, book_name: str = 'Workbook', folder_name: str = 'results'):
        self.book_name = book_name
        self.folder_name = folder_name
        self.writer = None
        self.sheet_name = None

    def __enter__(self):
        return self.open()

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()

    def open(self):
        if self.writer is not None:
            raise RuntimeError('Excel writer is already opened')

        book_path = get_state().get_file(self.folder_name, f'{self.book_name}.xlsx')
        self.writer = pd.ExcelWriter(book_path, engine='openpyxl', mode='w')

        return self

    def close(self):
        self.auto_width()
        # self.sort_sheets()
        self.writer.close()

    @property
    def workbook(self):
        return self.writer.book

    @property
    def worksheet(self):
        return self.writer.sheets[self.sheet_name]

    @worksheet.setter
    def worksheet(self, sheet_name: str):
        self.sheet_name = sheet_name
        if self.sheet_name not in self.writer.sheets:
            self.writer.book.create_sheet(self.sheet_name)

    def write_dataframe(self,
                        df: pd.DataFrame,
                        startrow: int = 0,
                        startcol: int = 0,
                        title: str = None,
                        color: bool = False,
                        invert_color: bool | list[bool] = False,
                        color_axis: str = 'both',
                        index: bool = True,
                        percentage: bool = False):
        df.to_excel(self.writer, sheet_name=self.sheet_name, startrow=startrow + 1, startcol=startcol, index=index)

        if title:
            cell = self.worksheet.cell(startrow + 1, startcol + 1, title)
            cell.font = Font(bold=True, color='FF0000')

            self.worksheet.merge_cells(start_row=startrow + 1,
                                       start_column=startcol + 1,
                                       end_row=startrow + 1,
                                       end_column=startcol + 1 + (df.shape[1] if index else df.shape[1] - 1))

        if color:
            # compute the bounds
            color_startrow = startrow + 3
            color_startcol = startcol + (2 if index else 1)
            color_endrow = startrow + 3 + df.shape[0]
            color_endcol = startcol + (2 if index else 1) + df.shape[1]

            # apply color formatting
            _format_color(self.worksheet, color_startrow, color_startcol, color_endrow, color_endcol, invert_color, color_axis)

        if percentage:
            # compute the bounds
            perc_startrow = startrow + 3
            perc_startcol = startcol + (2 if index else 1)
            perc_endrow = startrow + 3 + df.shape[0]
            perc_endcol = startcol + (2 if index else 1) + df.shape[1]

            # apply number formatting
            _format_percentage(self.worksheet, perc_startrow, perc_startcol, perc_endrow, perc_endcol)

    def write_text(self,
                   text: str | Number,
                   bold: bool = False,
                   align: str = None,
                   startrow: int = 0,
                   startcol: int = 0,
                   rowspan: int = 1,
                   colspan: int = 1):
        cell = self.worksheet.cell(startrow + 1, startcol + 1, text)

        if bold:
            cell.font = Font(bold=True)

        if align is not None:
            cell.alignment = Alignment(horizontal=align)

        if rowspan > 1 or colspan > 1:
            self.worksheet.merge_cells(start_row=startrow + 1,
                                       start_column=startcol + 1,
                                       end_row=startrow + 1 + (rowspan - 1),
                                       end_column=startcol + 1 + (colspan - 1))

    def write_image(self,
                    startrow: int = 0,
                    startcol: int = 0):

        stream = BytesIO()
        plt.savefig(stream, format='png')

        image = Image(stream)
        image.width = image.width // 2
        image.height = image.height // 2

        anchor = self.worksheet.cell(startrow + 1, startcol + 1).coordinate
        self.worksheet.add_image(image, anchor)

    def auto_width(self):
        for sheet in self.workbook._sheets:
            dim_holder = DimensionHolder(worksheet=sheet)

            for col in range(sheet.min_column, sheet.max_column + 1):
                col_letter = get_column_letter(col)

                width = 10
                for cell in sheet[col_letter]:
                    width = max(width, _get_cell_width(cell))
                width = min(width, 25)

                dim_holder[col_letter] = ColumnDimension(sheet, min=col, max=col, width=width)

            sheet.column_dimensions = dim_holder

    def sort_sheets(self):
        # sort by sheet title
        self.workbook._sheets.sort(key=lambda ws: ws.title)

        # pin summary sheet to the beginning
        if 'Summary' in self.workbook.sheetnames:
            self.workbook._sheets.insert(0, self.workbook._sheets.pop(self.workbook.sheetnames.index('Summary')))


def _get_cell_width(cell: Cell):
    value = cell.value
    columns = 1

    # check if merged cell
    sheet = cell.parent
    child_coord = cell.coordinate
    for merged in sheet.merged_cells.ranges:
        if child_coord in merged:
            value = merged.start_cell.value
            columns = merged.size['columns']

    # compute the width
    if isinstance(value, Number):
        return len(str(value)) // (2 * columns)
    else:
        if str(value).startswith("="):
            return 0
        return len(str(value)) // columns


def _format_color(worksheet: Worksheet,
                  color_startrow: int,
                  color_startcol: int,
                  color_endrow: int,
                  color_endcol: int,
                  invert_color: bool | list[bool] = False,
                  color_axis: str = 'both'):
    # color according to both axes
    if color_axis == 'both':
        if type(invert_color) is not bool:
            raise TypeError('invert_color cannot be a list if color_axis is "both"')

        color_scale_rule = _get_color_scale(invert_color)
        start_cell = f'{get_column_letter(color_startcol)}{color_startrow}'
        end_cell = f'{get_column_letter(color_endcol - 1)}{color_endrow - 1}'
        worksheet.conditional_formatting.add(f'{start_cell}:{end_cell}', color_scale_rule)

    # row-wise colors
    elif color_axis == 'row':
        if type(invert_color) is bool:
            invert_color = [invert_color for row in range(color_startrow, color_endrow)]

        for index, row in enumerate(range(color_startrow, color_endrow)):
            color_scale_rule = _get_color_scale(invert_color[index])
            start_cell = f'{get_column_letter(color_startcol)}{row}'
            end_cell = f'{get_column_letter(color_endcol - 1)}{row}'
            worksheet.conditional_formatting.add(f'{start_cell}:{end_cell}', color_scale_rule)

    # column-wise colors
    elif color_axis == 'column':
        if type(invert_color) is bool:
            invert_color = [invert_color for col in range(color_startcol, color_endcol)]

        for index, column in enumerate(range(color_startcol, color_endcol)):
            color_scale_rule = _get_color_scale(invert_color[index])
            start_cell = f'{get_column_letter(column)}{color_startrow}'
            end_cell = f'{get_column_letter(column)}{color_endrow - 1}'
            worksheet.conditional_formatting.add(f'{start_cell}:{end_cell}', color_scale_rule)


def _format_percentage(worksheet: Worksheet,
                       perc_startrow: int,
                       perc_startcol: int,
                       perc_endrow: int,
                       perc_endcol: int):
    for i in range(perc_startrow, perc_endrow):
        for j in range(perc_startcol, perc_endcol):
            worksheet[f'{get_column_letter(j)}{i}'].number_format = numbers.FORMAT_PERCENTAGE


def _get_color_scale(invert_color: bool) -> ColorScaleRule:
    if invert_color:
        return ColorScaleRule(start_type='percentile', start_value=10, start_color='FF63BE7B',
                              mid_type='num', mid_value=0, mid_color='FFFFFFFF',
                              end_type='percentile', end_value=90, end_color='FFF8696B')

    else:
        return ColorScaleRule(start_type='percentile', start_value=10, start_color='FFF8696B',
                              mid_type='num', mid_value=0, mid_color='FFFFFFFF',
                              end_type='percentile', end_value=90, end_color='FF63BE7B')
