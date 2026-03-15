from copy import copy

import openpyxl
from openpyxl.drawing.image import Image

from skylines.common.state import  get_state


def _copy_sheet(source_sheet, target_sheet):
    # Copy cell values and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]
            target_cell.value = cell.value

            # Copy cell style
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

    # Copy conditional formatting
    for cf, rules in source_sheet.conditional_formatting._cf_rules.items():
        for rule in rules:
            target_sheet.conditional_formatting.add(cf, rule)

    # Copy images
    for image in source_sheet._images:  # Access images directly from the source sheet
        img = Image(image.ref)  # Recreate the image
        target_sheet.add_image(img, image.anchor)  # Use the same anchor for placement


def merge_workbooks(out_file: str, in_files: list[str]):
    merged_workbook = openpyxl.Workbook()
    merged_workbook.remove(merged_workbook.active)  # Remove default empty sheet

    for in_file in in_files:
        in_file = get_state().get_file('results/intermediate', f'{in_file}.xlsx')
        workbook = openpyxl.load_workbook(in_file)

        for sheet_name in workbook.sheetnames:
            source_sheet = workbook[sheet_name]
            new_sheet = merged_workbook.create_sheet(sheet_name)

            # Copy data, formatting, and images
            _copy_sheet(source_sheet, new_sheet)

            # Optionally, adjust column widths
            for col in source_sheet.column_dimensions:
                new_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width

    # Save the merged workbook
    out_file = get_state().get_file('results', f'{out_file}.xlsx')
    merged_workbook.save(out_file)
