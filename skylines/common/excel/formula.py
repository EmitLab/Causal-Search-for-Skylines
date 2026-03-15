from openpyxl.utils import get_column_letter


def sum_if(criteria_range: tuple[int, int, int, int], criteria: tuple[int, int], sum_range: tuple[int, int, int, int]):
    return '_xlfn.SUMIF(' + \
        f'${get_column_letter(criteria_range[1])}${criteria_range[0]}:${get_column_letter(criteria_range[3])}${criteria_range[2]}' + ',' + \
        f'"*"&{get_column_letter(criteria[1])}{criteria[0]}&"*"' + ',' + \
        f'${get_column_letter(sum_range[1])}${sum_range[0]}:${get_column_letter(sum_range[3])}${sum_range[2]}' + ')'


def sum_ifs(sum_range: tuple[int, int, int, int], criteria_range_1: tuple[int, int, int, int],
            criteria_1: tuple[int, int], criteria_range_2: tuple[int, int, int, int], criteria_2: str):
    return '_xlfn.SUMIFS(' + \
        f'${get_column_letter(sum_range[1])}${sum_range[0]}:${get_column_letter(sum_range[3])}${sum_range[2]}' + ',' + \
        f'${get_column_letter(criteria_range_1[1])}${criteria_range_1[0]}:${get_column_letter(criteria_range_1[3])}${criteria_range_1[2]}' + ',' + \
        f'"*"&{get_column_letter(criteria_1[1])}{criteria_1[0]}&"*"' + ',' + \
        f'${get_column_letter(criteria_range_2[1])}${criteria_range_2[0]}:${get_column_letter(criteria_range_2[3])}${criteria_range_2[2]}' + ',' + \
        f'"{criteria_2}"' + ')'


def count_if(criteria_range: tuple[int, int, int, int], criteria: tuple[int, int]):
    return '_xlfn.COUNTIF(' + \
        f'${get_column_letter(criteria_range[1])}${criteria_range[0]}:${get_column_letter(criteria_range[3])}${criteria_range[2]}' + ',' + \
        f'"*"&{get_column_letter(criteria[1])}{criteria[0]}&"*"' + ')'


def max_ifs(max_range: tuple[int, int, int, int], criteria_range: tuple[int, int, int, int], criteria: tuple[int, int]):
    return '_xlfn.MAXIFS(' + \
        f'${get_column_letter(max_range[1])}${max_range[0]}:${get_column_letter(max_range[3])}${max_range[2]}' + ',' + \
        f'${get_column_letter(criteria_range[1])}${criteria_range[0]}:${get_column_letter(criteria_range[3])}${criteria_range[2]}' + ',' + \
        f'"*"&{get_column_letter(criteria[1])}{criteria[0]}&"*"' + ')'


def min_ifs(min_range: tuple[int, int, int, int], criteria_range: tuple[int, int, int, int], criteria: tuple[int, int]):
    return '_xlfn.MINIFS(' + \
        f'${get_column_letter(min_range[1])}${min_range[0]}:${get_column_letter(min_range[3])}${min_range[2]}' + ',' + \
        f'${get_column_letter(criteria_range[1])}${criteria_range[0]}:${get_column_letter(criteria_range[3])}${criteria_range[2]}' + ',' + \
        f'"*"&{get_column_letter(criteria[1])}{criteria[0]}&"*"' + ')'


def check_presence(criteria: tuple[int, int], cell: tuple[int, int]):
    return '_xlfn.IF(_xlfn.ISNUMBER(_xlfn.SEARCH(' + \
        f'"*"&{get_column_letter(criteria[1])}{criteria[0]}&"*"' + ',' + \
        f'${get_column_letter(cell[1])}${cell[0]})),"Yes", "No")'


def get_cell(cell: tuple[int, int]):
    return f'{get_column_letter(cell[1])}{cell[0]}'
